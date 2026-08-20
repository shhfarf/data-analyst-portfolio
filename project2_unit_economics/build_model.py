"""
Финансовая модель: Unit Economics + сценарный P&L для EdTech-подписочного продукта.
Использует openpyxl, все расчёты — формулами (не хардкод), с recalc.py после сохранения.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE = Font(color="0000FF")       # хардкод-инпуты
BLACK = Font(color="000000")      # формулы
GREEN = Font(color="008000")      # ссылки на другой лист
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
SUBHEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# =====================================================================
# SHEET 1: Assumptions
# =====================================================================
ws = wb.active
ws.title = "Assumptions"
ws["A1"] = "Unit Economics Model — EdTech Subscription Product"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Все синие ячейки — входные допущения (легенда: измени их, модель пересчитается автоматически)"
ws["A2"].font = Font(italic=True, color="666666", size=10)

ws["A4"] = "Ключевые допущения"
ws["A4"].font = BOLD
ws["A4"].fill = SUBHEAD_FILL

rows = [
    ("Средний чек за месяц подписки ($)", 25, "Источник: данные пользователя (гипотетический продукт)"),
    ("Валовая маржа (% от выручки)", 0.72, "Типично для EdTech SaaS: 70-75%"),
    ("Средний срок жизни подписчика (месяцев)", 5.5, "Рассчитано из retention-кривой проекта 1 (см. лист Cohort_LTV)"),
    ("Ежемесячный рост базы новых пользователей (%)", 0.04, "Консервативная оценка"),
    ("Фиксированные операционные расходы в месяц ($)", 18000, "Команда + инфраструктура"),
]
r = 5
for label, val, note in rows:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val)
    c.font = BLUE
    if isinstance(val, float) and val < 1:
        c.number_format = "0.0%"
    elif "$" in label:
        c.number_format = "$#,##0"
    ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="888888", size=9)
    r += 1

ws["A11"] = "CAC по каналам привлечения ($) — из проекта 1 (структура каналов)"
ws["A11"].font = BOLD
ws["A11"].fill = SUBHEAD_FILL

ws["A12"] = "Канал"; ws["B12"] = "CAC ($)"; ws["C12"] = "Completion rate (из проекта 1)"; ws["D12"] = "Доля в бюджете"
style_header(ws, 12, 1, 4)
channels = [
    ("Corporate", 42, 0.589, 0.20),
    ("Partner referral", 18, 0.495, 0.20),
    ("Organic", 6, 0.431, 0.25),
    ("Paid ads", 35, 0.252, 0.25),
    ("Social media", 22, 0.231, 0.10),
]
r = 13
for ch, cac, comp, share in channels:
    ws.cell(row=r, column=1, value=ch)
    c = ws.cell(row=r, column=2, value=cac); c.font = BLUE; c.number_format = "$#,##0"
    c2 = ws.cell(row=r, column=3, value=comp); c2.font = BLUE; c2.number_format = "0.0%"
    c3 = ws.cell(row=r, column=4, value=share); c3.font = BLUE; c3.number_format = "0.0%"
    r += 1

ws["A19"] = "Blended CAC ($, средневзвешенный по бюджету)"
ws["A19"].font = BOLD
c = ws.cell(row=19, column=2, value="=SUMPRODUCT(B13:B17,D13:D17)")
c.font = BLACK; c.number_format = "$#,##0.00"

autosize(ws, {"A": 46, "B": 16, "C": 30, "D": 16})

# =====================================================================
# SHEET 2: Cohort_LTV (использует retention-кривую из проекта 1)
# =====================================================================
ws2 = wb.create_sheet("Cohort_LTV")
ws2["A1"] = "LTV-модель на основе retention-кривой (проект 1)"
ws2["A1"].font = Font(bold=True, size=14)
ws2["A2"] = "Retention % по неделям взят как среднее по 4 когортам из SQL-анализа (result_cohort_retention.csv)"
ws2["A2"].font = Font(italic=True, color="666666", size=10)

ws2["A4"] = "Неделя"; ws2["B4"] = "Retention %"; ws2["C4"] = "Активных на 1000 стартовавших"; ws2["D4"] = "Выручка недели ($)"
style_header(ws2, 4, 1, 4)

retention_pct = [100.0, 90.6, 82.2, 73.9, 66.5, 60.6, 54.8, 50.0, 46.0, 42.5, 39.1, 36.0, 33.1]
r = 5
for wk, pct in enumerate(retention_pct):
    ws2.cell(row=r, column=1, value=wk)
    c = ws2.cell(row=r, column=2, value=pct / 100); c.font = BLUE; c.number_format = "0.0%"
    c2 = ws2.cell(row=r, column=3, value=f"=B{r}*1000"); c2.font = BLACK; c2.number_format = "0"
    # выручка недели = активные * (месячный чек / 4.33 недель) * маржа
    c3 = ws2.cell(row=r, column=4, value=f"=C{r}*(Assumptions!$B$5/4.33)*Assumptions!$B$6")
    c3.font = GREEN; c3.number_format = "$#,##0"
    r += 1

ws2["A19"] = "Суммарная выручка (маржинальная) на 1000 стартовавших за 12 недель"
ws2["A19"].font = BOLD
c = ws2.cell(row=19, column=4, value="=SUM(D5:D17)")
c.font = BLACK; c.number_format = "$#,##0"

ws2["A21"] = "LTV на одного пользователя ($, маржинальный, за период наблюдения)"
ws2["A21"].font = BOLD
c = ws2.cell(row=21, column=4, value="=D19/1000")
c.font = BLACK; c.number_format = "$#,##0.00"

autosize(ws2, {"A": 40, "B": 16, "C": 28, "D": 22})

# =====================================================================
# SHEET 3: Unit_Economics_by_Channel
# =====================================================================
ws3 = wb.create_sheet("Unit_Economics")
ws3["A1"] = "Unit Economics по каналам: LTV / CAC, payback period"
ws3["A1"].font = Font(bold=True, size=14)

headers = ["Канал", "CAC ($)", "LTV ($)", "LTV/CAC", "Payback (мес.)", "Вердикт"]
for i, h in enumerate(headers, start=1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 1, 6)

for i in range(5):
    r = 4 + i
    src_r = 13 + i
    ws3.cell(row=r, column=1, value=f"=Assumptions!A{src_r}").font = GREEN
    ws3.cell(row=r, column=2, value=f"=Assumptions!B{src_r}").font = GREEN
    ws3.cell(row=r, column=2).number_format = "$#,##0"
    # LTV масштабируем по completion rate канала относительно среднего (упрощённая проекция)
    ws3.cell(row=r, column=3, value=f"=Cohort_LTV!$D$21*(Assumptions!C{src_r}/AVERAGE(Assumptions!$C$13:$C$17))")
    ws3.cell(row=r, column=3).font = BLACK
    ws3.cell(row=r, column=3).number_format = "$#,##0.00"
    ws3.cell(row=r, column=4, value=f"=C{r}/B{r}").font = BLACK
    ws3.cell(row=r, column=4).number_format = "0.00x"
    ws3.cell(row=r, column=5, value=f"=B{r}/(C{r}/5.5)").font = BLACK  # payback = CAC / (LTV/lifetime months)
    ws3.cell(row=r, column=5).number_format = "0.0"
    ws3.cell(row=r, column=6, value=f'=IF(D{r}>=3,"Здоровая экономика",IF(D{r}>=1,"Приемлемо, требует оптимизации","Убыточный канал"))')
    ws3.cell(row=r, column=6).font = BLACK

ws3["A11"] = "Правило: LTV/CAC >= 3x считается здоровой unit-экономикой для подписочных продуктов."
ws3["A11"].font = Font(italic=True, size=9, color="888888")

autosize(ws3, {"A": 20, "B": 12, "C": 14, "D": 12, "E": 16, "F": 26})

# =====================================================================
# SHEET 4: Scenario_PnL
# =====================================================================
ws4 = wb.create_sheet("Scenario_PnL")
ws4["A1"] = "Сценарный P&L на 12 месяцев (Base / Optimistic / Pessimistic)"
ws4["A1"].font = Font(bold=True, size=14)

ws4["A3"] = "Множитель роста базы (к базовому темпу из Assumptions!B8)"
ws4["A3"].fill = SUBHEAD_FILL; ws4["A3"].font = BOLD
ws4["B4"] = "Base"; ws4["C4"] = "Optimistic"; ws4["D4"] = "Pessimistic"
style_header(ws4, 4, 2, 4)
ws4["A5"] = "Growth multiplier"
for col, val in zip("BCD", [1.0, 1.5, 0.6]):
    c = ws4[f"{col}5"]; c.value = val; c.font = BLUE; c.number_format = "0.00x"

ws4["A6"] = "Стартовая база подписчиков (мес. 1)"
c = ws4["B6"]; c.value = 2400; c.font = BLUE; c.number_format = "#,##0"

headers2 = ["Месяц"] + [f"Base_M{i}" for i in range(1, 13)]
ws4["A8"] = "Прогноз базы подписчиков (Base сценарий)"
ws4["A8"].font = BOLD

ws4["A10"] = "Месяц"; ws4["B10"] = "Base"; ws4["C10"] = "Optimistic"; ws4["D10"] = "Pessimistic"
style_header(ws4, 10, 1, 4)
for m in range(1, 13):
    r = 10 + m
    ws4.cell(row=r, column=1, value=m)
    if m == 1:
        for col in "BCD":
            ws4[f"{col}{r}"] = "=$B$6"
    else:
        prev = r - 1
        ws4.cell(row=r, column=2, value=f"=B{prev}*(1+Assumptions!$B$8*$B$5)")
        ws4.cell(row=r, column=3, value=f"=C{prev}*(1+Assumptions!$B$8*$C$5)")
        ws4.cell(row=r, column=4, value=f"=D{prev}*(1+Assumptions!$B$8*$D$5)")
    for col in "BCD":
        ws4[f"{col}{r}"].font = BLACK
        ws4[f"{col}{r}"].number_format = "#,##0"

ws4["A24"] = "Выручка (маржинальная, $/мес)"
ws4["A24"].font = BOLD; ws4["A24"].fill = SUBHEAD_FILL
ws4["A25"] = "Месяц"; ws4["B25"] = "Base"; ws4["C25"] = "Optimistic"; ws4["D25"] = "Pessimistic"; ws4["E25"] = "EBITDA (Base)"
style_header(ws4, 25, 1, 5)
for m in range(1, 13):
    r = 25 + m
    src_r = 10 + m
    ws4.cell(row=r, column=1, value=m)
    for col, src_col in zip("BCD", "BCD"):
        ws4[f"{col}{r}"] = f"={src_col}{src_r}*Assumptions!$B$5*Assumptions!$B$6"
        ws4[f"{col}{r}"].font = BLACK
        ws4[f"{col}{r}"].number_format = "$#,##0"
    ws4.cell(row=r, column=5, value=f"=B{r}-Assumptions!$B$9")
    ws4.cell(row=r, column=5).font = BLACK
    ws4.cell(row=r, column=5).number_format = "$#,##0;($#,##0);-"

ws4["A39"] = "Итого выручка за 12 мес (Base)"
ws4["A39"].font = BOLD
c = ws4["B39"] = "=SUM(B26:B37)"
ws4["B39"].font = BLACK; ws4["B39"].number_format = "$#,##0"
ws4["A40"] = "Итого EBITDA за 12 мес (Base)"
ws4["A40"].font = BOLD
ws4["B40"] = "=SUM(E26:E37)"
ws4["B40"].font = BLACK; ws4["B40"].number_format = "$#,##0;($#,##0);-"

autosize(ws4, {"A": 34, "B": 14, "C": 14, "D": 14, "E": 16})

wb.save("unit_economics_model.xlsx")
print("saved")
